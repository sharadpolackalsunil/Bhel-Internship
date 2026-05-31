"""
Data Processing & Export Module
=================================
Takes scraped result data and processes it into:
    1. results.csv — All students sorted by SGPA (ascending)
    2. results.xlsx — Excel workbook with:
        - 'All Students' sheet (combined, sorted)
        - 'AI_DS' sheet (BTAD students)
        - 'AI_ML' sheet (BTAM students)
        - 'AI' sheet (BTAI students)

Extracts: enrollment, name, branch, semester, SGPA, CGPA,
          result status (pass/fail), grades, and subject details.
"""

import os
import sys
import json

import pandas as pd
from datetime import datetime

# Add project root to path
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from scraper.enrollment import BRANCHES


def load_raw_results(raw_path=None):
    """
    Load raw scraped results from JSON file.

    Args:
        raw_path: Path to raw JSON file. If None, uses default location.

    Returns:
        List of result dictionaries.
    """
    if raw_path is None:
        raw_path = os.path.join(project_root, 'data', 'raw',
                                'raw_results_4.json')

    if not os.path.exists(raw_path):
        # Try checkpoint file
        checkpoint_path = os.path.join(project_root, 'data',
                                       'scrape_checkpoint.json')
        if os.path.exists(checkpoint_path):
            raw_path = checkpoint_path
        else:
            raise FileNotFoundError(
                f"No results file found at: {raw_path}\n"
                "Run the scraper first: python -m scraper.scraper"
            )

    with open(raw_path, 'r', encoding='utf-8') as f:
        results = json.load(f)

    print(f"📂 Loaded {len(results)} results from: {raw_path}")
    return results


def clean_sgpa(value):
    """Clean and parse SGPA value to float."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    # Remove whitespace and common artifacts
    value = str(value).strip()
    value = value.replace(',', '.')

    try:
        sgpa = float(value)
        if 0.0 <= sgpa <= 10.0:
            return sgpa
    except (ValueError, TypeError):
        pass

    return None


def process_results(results):
    """
    Process raw scraped results into a clean DataFrame.

    Args:
        results: List of result dictionaries from the scraper.

    Returns:
        pd.DataFrame: Cleaned and processed data.
    """
    print("\n🔄 Processing results...")

    # Build rows
    rows = []
    for r in results:
        enrollment = r.get('enrollment', 'UNKNOWN')
        status = r.get('result_status', 'UNKNOWN')

        # Skip completely failed scrapes
        if status == 'SCRAPE_FAILED':
            continue

        row = {
            'Enrollment_No': enrollment,
            'Student_Name': r.get('student_name', '—'),
            'Father_Name': r.get('father_name', '—'),
            'Branch': r.get('branch', '—'),
            'Branch_Code': r.get('branch_short', '—'),
            'Semester': r.get('semester', '4'),
            'Program': r.get('program', 'B.Tech'),
            'SGPA': clean_sgpa(r.get('sgpa')),
            'CGPA': clean_sgpa(r.get('cgpa')),
            'Result_Status': status.upper() if status else '—',
            'Total_Marks': r.get('total_marks', '—'),
            'Max_Marks': r.get('max_marks', '—'),
        }

        # Add subject-wise data as additional columns
        subjects = r.get('subjects', [])
        for j, subj in enumerate(subjects):
            for col_key, col_val in subj.items():
                row[f'Subject_{j+1}_{col_key}'] = col_val

        rows.append(row)

    df = pd.DataFrame(rows)

    # Clean up
    if 'SGPA' in df.columns:
        df['SGPA'] = pd.to_numeric(df['SGPA'], errors='coerce')

    if 'CGPA' in df.columns:
        df['CGPA'] = pd.to_numeric(df['CGPA'], errors='coerce')

    # Sort by SGPA in ascending order (user requested ascending)
    df = df.sort_values(by='SGPA', ascending=True, na_position='last')

    # Add rank (ascending SGPA means rank 1 = lowest, but let's rank desc)
    # Actually, rank by SGPA descending (best students get rank 1)
    df['Rank'] = df['SGPA'].rank(ascending=False, method='min').astype('Int64')

    # Reset index
    df = df.reset_index(drop=True)

    print(f"  ✅ Processed {len(df)} student records")
    print(f"  📊 SGPA range: {df['SGPA'].min():.2f} — {df['SGPA'].max():.2f}"
          if df['SGPA'].notna().any() else "  ⚠️  No SGPA data available")

    return df


def export_to_csv(df, output_path=None):
    """
    Export DataFrame to CSV file.

    Args:
        df: Processed DataFrame.
        output_path: Output file path. Defaults to data/results.csv.

    Returns:
        str: Path to saved file.
    """
    if output_path is None:
        output_path = os.path.join(project_root, 'data', 'results.csv')

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df.to_csv(output_path, index=False, encoding='utf-8-sig')
    print(f"📄 CSV saved: {output_path} ({len(df)} rows)")
    return output_path


def export_to_excel(df, output_path=None):
    """
    Export DataFrame to Excel with separate sheets per branch.

    Sheets:
        - All Students: Combined view sorted by SGPA (ascending)
        - AI_DS: BTAD students
        - AI_ML: BTAM students
        - AI: BTAI students

    Args:
        df: Processed DataFrame.
        output_path: Output file path. Defaults to data/results.xlsx.

    Returns:
        str: Path to saved file.
    """
    if output_path is None:
        output_path = os.path.join(project_root, 'data', 'results.xlsx')

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: All students
        df.to_excel(writer, sheet_name='All Students', index=False)

        # Format the "All Students" sheet
        worksheet = writer.sheets['All Students']
        _format_excel_sheet(worksheet, df)

        # Branch-specific sheets
        for code, info in BRANCHES.items():
            branch_df = df[df['Branch_Code'] == info['short_name']].copy()
            branch_df = branch_df.sort_values(
                by='SGPA', ascending=True, na_position='last'
            )
            branch_df = branch_df.reset_index(drop=True)

            # Add branch-specific rank
            branch_df['Branch_Rank'] = branch_df['SGPA'].rank(
                ascending=False, method='min'
            ).astype('Int64')

            sheet_name = info['short_name']
            branch_df.to_excel(writer, sheet_name=sheet_name, index=False)

            worksheet = writer.sheets[sheet_name]
            _format_excel_sheet(worksheet, branch_df)

            print(f"  📊 Sheet '{sheet_name}': {len(branch_df)} students")

    print(f"📊 Excel saved: {output_path}")
    return output_path


def _format_excel_sheet(worksheet, df):
    """Apply formatting to an Excel worksheet."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Header styling
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79',
                              fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Apply header formatting
    for col_idx in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Auto-adjust column widths
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            len(str(worksheet.cell(row=1, column=col_idx).value or '')),
            max((len(str(worksheet.cell(row=r, column=col_idx).value or ''))
                 for r in range(2, min(worksheet.max_row + 1, 50))),
                default=0)
        )
        adjusted_width = min(max_length + 4, 40)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    # Apply data cell formatting
    data_alignment = Alignment(horizontal='center', vertical='center')
    for row in range(2, worksheet.max_row + 1):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = data_alignment

            # Highlight SGPA column
            col_name = df.columns[col - 1] if col <= len(df.columns) else ''
            if col_name == 'SGPA' and cell.value is not None:
                try:
                    sgpa = float(cell.value)
                    if sgpa >= 9.0:
                        cell.fill = PatternFill(start_color='C6EFCE',
                                                end_color='C6EFCE',
                                                fill_type='solid')
                    elif sgpa >= 7.5:
                        cell.fill = PatternFill(start_color='FFEB9C',
                                                end_color='FFEB9C',
                                                fill_type='solid')
                    elif sgpa < 5.0:
                        cell.fill = PatternFill(start_color='FFC7CE',
                                                end_color='FFC7CE',
                                                fill_type='solid')
                except (ValueError, TypeError):
                    pass

            # Color-code Pass/Fail
            if col_name == 'Result_Status':
                text = str(cell.value or '').upper()
                if text == 'PASS':
                    cell.font = Font(color='006100', bold=True)
                    cell.fill = PatternFill(start_color='C6EFCE',
                                            end_color='C6EFCE',
                                            fill_type='solid')
                elif text == 'FAIL':
                    cell.font = Font(color='9C0006', bold=True)
                    cell.fill = PatternFill(start_color='FFC7CE',
                                            end_color='FFC7CE',
                                            fill_type='solid')

    # Freeze top row
    worksheet.freeze_panes = 'A2'

    # Add auto-filter
    worksheet.auto_filter.ref = worksheet.dimensions


def generate_summary_report(df):
    """
    Print a summary report of the results.

    Args:
        df: Processed DataFrame.
    """
    print("\n" + "=" * 60)
    print("📊 RESULTS SUMMARY REPORT")
    print(f"   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    print(f"\n  Total Students:  {len(df)}")
    print(f"  Semester:        4")

    if df['SGPA'].notna().any():
        print(f"\n  SGPA Statistics (All Branches):")
        print(f"    Mean:    {df['SGPA'].mean():.2f}")
        print(f"    Median:  {df['SGPA'].median():.2f}")
        print(f"    Std Dev: {df['SGPA'].std():.2f}")
        print(f"    Min:     {df['SGPA'].min():.2f}")
        print(f"    Max:     {df['SGPA'].max():.2f}")

    # Per-branch stats
    print(f"\n  Per-Branch Breakdown:")
    print(f"  {'Branch':<25} {'Count':>6} {'Avg SGPA':>10} {'Top SGPA':>10} {'Pass%':>8}")
    print(f"  {'-'*25} {'-'*6} {'-'*10} {'-'*10} {'-'*8}")

    for code, info in BRANCHES.items():
        branch_df = df[df['Branch_Code'] == info['short_name']]
        count = len(branch_df)

        if count > 0 and branch_df['SGPA'].notna().any():
            avg = branch_df['SGPA'].mean()
            top = branch_df['SGPA'].max()
            pass_count = len(branch_df[
                branch_df['Result_Status'].str.upper() == 'PASS'
            ])
            pass_pct = (pass_count / count * 100) if count > 0 else 0
            print(f"  {info['name']:<25} {count:>6} {avg:>10.2f} {top:>10.2f} "
                  f"{pass_pct:>7.1f}%")
        else:
            print(f"  {info['name']:<25} {count:>6} {'—':>10} {'—':>10} {'—':>8}")

    # Top students
    if df['SGPA'].notna().any():
        print(f"\n  🏆 Top 10 Students (by SGPA):")
        top_10 = df.nlargest(10, 'SGPA')[
            ['Enrollment_No', 'Student_Name', 'Branch_Code', 'SGPA']
        ]
        for idx, row in top_10.iterrows():
            print(f"    {row['Enrollment_No']}  {row['Student_Name']:<20} "
                  f"{row['Branch_Code']:<6} SGPA: {row['SGPA']:.2f}")

    print("\n" + "=" * 60)


def run_export(raw_path=None, csv_path=None, xlsx_path=None):
    """
    Complete export pipeline: load → process → export.

    Args:
        raw_path: Path to raw results JSON.
        csv_path: Output CSV path.
        xlsx_path: Output Excel path.

    Returns:
        pd.DataFrame: The processed DataFrame.
    """
    # Load raw data
    results = load_raw_results(raw_path)

    # Process
    df = process_results(results)

    # Export
    export_to_csv(df, csv_path)
    export_to_excel(df, xlsx_path)

    # Summary
    generate_summary_report(df)

    return df


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Export Results to CSV/Excel")
    parser.add_argument('--input', type=str, default=None,
                        help='Path to raw results JSON')
    parser.add_argument('--csv', type=str, default=None,
                        help='Output CSV path')
    parser.add_argument('--xlsx', type=str, default=None,
                        help='Output Excel path')
    args = parser.parse_args()

    run_export(raw_path=args.input, csv_path=args.csv, xlsx_path=args.xlsx)
